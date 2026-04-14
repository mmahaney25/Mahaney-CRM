#!/usr/bin/env python3
"""Create the Email_Outreach_Manager.xlsx workbook with proper structure.

Generates the workbook with two sheets matching the exact layout expected
by the MAHANEY CRM Python tool (see lib/excel_io.py for column constants).

Sheets:
  - "Outreach Tracker" — master contact view with sync statuses
  - "Outreach" — action page for composing and sending messages
"""

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = Path(os.path.expanduser("~/Downloads/Email_Outreach_Manager.xlsx"))

# --- Tracker sheet (matches TR_* constants in lib/excel_io.py) ---
# A=Name, B=General Status, C=Priority, D=Age,
# E=Email Status, F=Text Status, G=LinkedIn Status,
# H=Email, I=Number
TRACKER_HEADERS = [
    ("Name", 22),
    ("General Status", 16),
    ("Priority", 12),
    ("Age", 8),
    ("Email Status", 15),
    ("Text Status", 15),
    ("LinkedIn Status", 17),
    ("Email", 30),
    ("Number", 18),
]

# --- Outreach sheet (matches OR_* constants in lib/excel_io.py) ---
# A=Name, B=Status, C=Email Contact, D=Messages Contact,
# E=Company, F=Message
OUTREACH_HEADERS = [
    ("Name", 22),
    ("Status", 12),
    ("Email Contact", 30),
    ("Messages Contact", 20),
    ("Company", 22),
    ("Message", 50),
]

# Styles
HEADER_FONT = Font(bold=True, size=11)
HEADER_BORDER = Border(bottom=Side(style="medium"))
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _write_headers(ws, headers):
    """Write header row and set column widths."""
    for col_idx, (header_text, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[cell.column_letter].width = width


def main():
    wb = Workbook()

    # --- Outreach Tracker sheet ---
    ws_tracker = wb.active
    ws_tracker.title = "Outreach Tracker"
    _write_headers(ws_tracker, TRACKER_HEADERS)

    # Priority dropdown validation (col C, rows 2-500)
    priority_validation = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True,
    )
    priority_validation.error = "Please select High, Medium, or Low"
    priority_validation.errorTitle = "Invalid Priority"
    ws_tracker.add_data_validation(priority_validation)
    priority_validation.add("C2:C500")

    # Freeze header row
    ws_tracker.freeze_panes = "A2"

    # --- Outreach sheet ---
    ws_outreach = wb.create_sheet("Outreach")
    _write_headers(ws_outreach, OUTREACH_HEADERS)

    # Freeze header row
    ws_outreach.freeze_panes = "A2"

    # --- Save ---
    wb.save(str(OUTPUT_PATH))
    print(f"Workbook created: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
