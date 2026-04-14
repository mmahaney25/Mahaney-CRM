"""Excel read/write helpers via Microsoft Graph Excel REST API.

Online mode (default): reads and writes cells directly through the Graph API —
no download/upload cycle.  Local mode (SPREADSHEET_PATH set): falls back to
openpyxl for offline development.
"""

import io
import os
import urllib.parse
from pathlib import Path

import requests

from lib.auth import get_token

# OneDrive path (relative to drive root)
ONEDRIVE_PATH = os.environ.get("ONEDRIVE_PATH", "Email_Outreach_Manager.xlsx")

# Local fallback — if set and the file exists, skip Graph API
_local = os.environ.get("SPREADSHEET_PATH", "")
LOCAL_PATH = Path(os.path.expanduser(_local)) if _local else None

GRAPH_BASE = "https://graph.microsoft.com/v1.0"

TRACKER_SHEET = "Outreach Tracker"
OUTREACH_SHEET = "Outreach"

# Tracker column indices (1-based) — matches actual spreadsheet layout
TR_NAME = 1        # A - Contact name
TR_GENERAL = 2     # B - Status (rollup: Reply/Sent)
TR_COMPANY = 3     # C - Company
TR_ROLE = 4        # D - Role
TR_PRIORITY = 5    # E - High/Medium/Low
TR_EMAIL = 6       # F - Email address
TR_PHONE = 7       # G - Phone number
TR_AGE = 8         # H - Age string (e.g. "14h", "4d")

# Outreach column indices (1-based)
OR_NAME = 1       # A
OR_STATUS = 2     # B
OR_EMAIL = 3      # C
OR_PHONE = 4      # D
OR_COMPANY = 5    # E
OR_MESSAGE = 6    # F

# Priority colors (used only in local/openpyxl mode)
try:
    from openpyxl.styles import PatternFill
    PRIORITY_FILLS = {
        "High": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "Medium": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "Low": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
    }
except ImportError:
    PRIORITY_FILLS = {}


def _use_local() -> bool:
    return LOCAL_PATH is not None and LOCAL_PATH.exists()


# ---------------------------------------------------------------------------
# Column helpers
# ---------------------------------------------------------------------------

def _col_letter(col: int) -> str:
    """Convert 1-based column index to Excel letter(s): 1→A, 26→Z, 27→AA."""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _cell_address(row: int, col: int) -> str:
    """Return Excel-style address like 'E2'."""
    return f"{_col_letter(col)}{row}"


# ---------------------------------------------------------------------------
# Graph API helpers (online mode)
# ---------------------------------------------------------------------------

def _graph_headers(*, content_type: str | None = None) -> dict:
    token = get_token(["Files.ReadWrite"])
    h = {"Authorization": f"Bearer {token}"}
    if content_type:
        h["Content-Type"] = content_type
    return h


def _refresh_graph_headers(*, content_type: str | None = None) -> dict:
    """Force a fresh token (clears the module-level cache) and return headers."""
    import lib.auth
    lib.auth._cached_token = None
    return _graph_headers(content_type=content_type)


def _sheet_url(sheet: str) -> str:
    encoded = urllib.parse.quote(sheet)
    return f"{GRAPH_BASE}/me/drive/root:/{ONEDRIVE_PATH}:/workbook/worksheets/{encoded}"


def _graph_request(method: str, url: str, *, content_type: str | None = None,
                   json_body: dict | None = None, timeout: int = 30) -> requests.Response:
    """Make a Graph API request with automatic retry on 401 (expired token)."""
    headers = _graph_headers(content_type=content_type)
    resp = _do_request(method, url, headers, json_body, timeout)

    if resp.status_code == 401:
        print("  [graph] Token expired, refreshing...", flush=True)
        headers = _refresh_graph_headers(content_type=content_type)
        resp = _do_request(method, url, headers, json_body, timeout)

    return resp


def _do_request(method, url, headers, json_body, timeout):
    if method == "GET":
        return requests.get(url, headers=headers, timeout=timeout)
    elif method == "PATCH":
        return requests.patch(url, headers=headers, json=json_body, timeout=timeout)
    raise ValueError(f"Unsupported method: {method}")


def read_range(sheet: str, address: str) -> list[list]:
    """Read a range and return a 2-D list of values."""
    url = f"{_sheet_url(sheet)}/range(address='{address}')"
    resp = _graph_request("GET", url)
    if resp.status_code != 200:
        raise RuntimeError(f"Graph read_range failed ({resp.status_code}): {resp.text[:300]}")
    return resp.json().get("values", [])


def read_used_range(sheet: str) -> list[list]:
    """Read the entire used range of a sheet (header + data rows)."""
    url = f"{_sheet_url(sheet)}/usedRange"
    resp = _graph_request("GET", url)
    if resp.status_code != 200:
        raise RuntimeError(f"Graph usedRange failed ({resp.status_code}): {resp.text[:300]}")
    return resp.json().get("values", [])


def write_cell(sheet: str, row: int, col: int, value) -> None:
    """Write a single cell value via Graph API."""
    addr = _cell_address(row, col)
    url = f"{_sheet_url(sheet)}/range(address='{addr}')"
    payload = {"values": [[value if value is not None else ""]]}
    resp = _graph_request("PATCH", url, content_type="application/json", json_body=payload)
    if resp.status_code != 200:
        raise RuntimeError(f"Graph write_cell {addr} failed ({resp.status_code}): {resp.text[:300]}")


def write_range(sheet: str, address: str, values: list[list]) -> None:
    """Write a 2-D block of values to a range."""
    url = f"{_sheet_url(sheet)}/range(address='{address}')"
    resp = _graph_request("PATCH", url, content_type="application/json",
                          json_body={"values": values}, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"Graph write_range {address} failed ({resp.status_code}): {resp.text[:300]}")


# ---------------------------------------------------------------------------
# High-level data accessors
# ---------------------------------------------------------------------------

def get_tracker_rows() -> list[dict]:
    """Read all contact rows from the Tracker sheet (row 2+).

    Returns list of dicts with keys: row, name, general_status, priority,
    age, email_status, text_status, linkedin_status, email, phone.
    """
    if _use_local():
        return _local_get_tracker_rows()

    all_values = read_used_range(TRACKER_SHEET)
    if len(all_values) < 2:
        return []

    rows = []
    for i, vals in enumerate(all_values[1:], start=2):
        # Pad short rows to at least 8 columns (A-H)
        while len(vals) < 8:
            vals.append(None)
        name = vals[TR_NAME - 1]
        if not name:
            continue
        rows.append({
            "row": i,
            "name": str(name).strip(),
            "general_status": vals[TR_GENERAL - 1],
            "priority": vals[TR_PRIORITY - 1],
            "age": vals[TR_AGE - 1],
            "email": str(vals[TR_EMAIL - 1] or "").strip().lower(),
            "phone": str(vals[TR_PHONE - 1] or "").strip() if vals[TR_PHONE - 1] else "",
            "company": vals[TR_COMPANY - 1],
            "role": vals[TR_ROLE - 1],
        })
    return rows


def get_outreach_rows() -> list[dict]:
    """Read all rows from the Outreach sheet (row 2+)."""
    if _use_local():
        return _local_get_outreach_rows()

    try:
        all_values = read_used_range(OUTREACH_SHEET)
    except RuntimeError:
        print(f"  WARNING: '{OUTREACH_SHEET}' sheet not found or unreadable. Skipping.")
        return []

    if len(all_values) < 2:
        return []

    rows = []
    for i, vals in enumerate(all_values[1:], start=2):
        while len(vals) < 6:
            vals.append(None)
        name = vals[0]
        if not name:
            continue
        rows.append({
            "row": i,
            "name": str(name).strip(),
            "status": vals[1],
            "email_contact": str(vals[2] or "").strip(),
            "messages_contact": str(vals[3] or "").strip(),
            "company": vals[4],
            "message": str(vals[5] or "").strip(),
        })
    return rows


# ---------------------------------------------------------------------------
# High-level writers
# ---------------------------------------------------------------------------

def write_tracker_cell(row: int, col: int, value) -> None:
    """Write a value to a specific Tracker cell."""
    if _use_local():
        raise RuntimeError("Local mode writes not supported via this function — use openpyxl directly")
    write_cell(TRACKER_SHEET, row, col, value)


def write_outreach_cell(row: int, col: int, value) -> None:
    """Write a value to a specific Outreach cell."""
    if _use_local():
        raise RuntimeError("Local mode writes not supported via this function — use openpyxl directly")
    write_cell(OUTREACH_SHEET, row, col, value)


def write_tracker_range(start_row: int, end_row: int, values: list[list]) -> None:
    """Write a block of rows to the Tracker sheet (for sorting).

    values is a 2-D list where each inner list is a full row of cell values.
    Writes from column A to the width of the first row.
    """
    if not values:
        return
    ncols = len(values[0])
    end_col = _col_letter(ncols)
    address = f"A{start_row}:{end_col}{end_row}"
    write_range(TRACKER_SHEET, address, values)


# ---------------------------------------------------------------------------
# Local fallback (openpyxl) — only used when SPREADSHEET_PATH is set
# ---------------------------------------------------------------------------

def _local_get_tracker_rows() -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(str(LOCAL_PATH))
    ws = wb[TRACKER_SHEET]
    rows = []
    for row_num in range(2, ws.max_row + 1):
        name = ws.cell(row=row_num, column=TR_NAME).value
        if not name:
            continue
        rows.append({
            "row": row_num,
            "name": str(name).strip(),
            "general_status": ws.cell(row=row_num, column=TR_GENERAL).value,
            "priority": ws.cell(row=row_num, column=TR_PRIORITY).value,
            "age": ws.cell(row=row_num, column=TR_AGE).value,
            "email": str(ws.cell(row=row_num, column=TR_EMAIL).value or "").strip().lower(),
            "phone": str(ws.cell(row=row_num, column=TR_PHONE).value or "").strip(),
            "company": ws.cell(row=row_num, column=TR_COMPANY).value,
            "role": ws.cell(row=row_num, column=TR_ROLE).value,
        })
    return rows


def _local_get_outreach_rows() -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(str(LOCAL_PATH))
    if OUTREACH_SHEET not in wb.sheetnames:
        print(f"  WARNING: '{OUTREACH_SHEET}' sheet not found in workbook. Skipping.")
        return []
    ws = wb[OUTREACH_SHEET]
    rows = []
    for row_num in range(2, ws.max_row + 1):
        name = ws.cell(row=row_num, column=OR_NAME).value
        if not name:
            continue
        rows.append({
            "row": row_num,
            "name": str(name).strip(),
            "status": ws.cell(row=row_num, column=OR_STATUS).value,
            "email_contact": str(ws.cell(row=row_num, column=OR_EMAIL).value or "").strip(),
            "messages_contact": str(ws.cell(row=row_num, column=OR_PHONE).value or "").strip(),
            "company": ws.cell(row=row_num, column=OR_COMPANY).value,
            "message": str(ws.cell(row=row_num, column=OR_MESSAGE).value or "").strip(),
        })
    return rows
